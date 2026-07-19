#!/usr/bin/env python3
"""Importiert Teilnehmer aus TN-Liste_Tenero 2026.xlsx in die Datenbank."""

import sqlite3
import os
import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), "daten", "daten.db")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "TN-Liste_Tenero 2026.xlsx")

print(f"📂 DB: {DB_PATH}")
print(f"📂 Excel: {XLSX_PATH}")

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row

# ─── Migration: Neue Spalten ───
print("\n🔧 Migration: Neue Spalten...")
cols = {row["name"] for row in conn.execute("PRAGMA table_info(teilnehmer)")}
migrations = [
    ("anrede", "TEXT NOT NULL DEFAULT ''"),
    ("kommentar", "TEXT NOT NULL DEFAULT ''"),
    ("schwimmen", "TEXT NOT NULL DEFAULT ''"),
    ("medikamente", "TEXT NOT NULL DEFAULT ''"),
    ("hauptsportart", "TEXT NOT NULL DEFAULT ''"),
]
for col, coltype in migrations:
    if col not in cols:
        conn.execute(f"ALTER TABLE teilnehmer ADD COLUMN {col} {coltype}")
        print(f"  ✅ Spalte '{col}' hinzugefügt")
    else:
        print(f"  ⏭️  Spalte '{col}' existiert bereits")

conn.commit()

# ─── Excel lesen ───
print("\n📖 Excel lesen...")
wb = openpyxl.load_workbook(XLSX_PATH)

# --- Zusatzformular (Hauptdaten) ---
ws = wb['Zusatzformular']
formular_data = {}
for r in range(6, ws.max_row + 1):
    anrede = (ws.cell(row=r, column=1).value or "").strip()
    nachname = (ws.cell(row=r, column=2).value or "").strip()
    vorname = (ws.cell(row=r, column=3).value or "").strip()
    hauptsportart = (ws.cell(row=r, column=5).value or "").strip()
    schwimmkenntnisse = (ws.cell(row=r, column=7).value or "").strip()
    allergien = (ws.cell(row=r, column=18).value or "").strip()
    medikamente = (ws.cell(row=r, column=21).value or "").strip()
    aufmerksamkeit = (ws.cell(row=r, column=22).value or "").strip()

    if not vorname or not nachname:
        continue

    # Kommentar = Allergien + Aufmerksamkeit
    kommentar_teile = []
    if allergien:
        kommentar_teile.append(f"Allergien: {allergien}")
    if aufmerksamkeit:
        kommentar_teile.append(f"Aufmerksamkeit: {aufmerksamkeit}")
    kommentar = " | ".join(kommentar_teile)

    key = (vorname, nachname)
    formular_data[key] = {
        "anrede": anrede,
        "vorname": vorname,
        "nachname": nachname,
        "hauptsportart": hauptsportart,
        "schwimmen": schwimmkenntnisse,
        "medikamente": medikamente,
        "kommentar": kommentar,
    }

print(f"  {len(formular_data)} Einträge aus Zusatzformular")

# --- Einteilung Zelte (Zelt-Zuordnung) ---
ws_z = wb['Einteilung Zelte']
zelt_data = {}
for r in range(6, ws_z.max_row + 1):
    anrede = (ws_z.cell(row=r, column=1).value or "").strip()
    vorname = (ws_z.cell(row=r, column=2).value or "").strip()
    name = (ws_z.cell(row=r, column=3).value or "").strip()
    zelt = ws_z.cell(row=r, column=6).value

    if not vorname or not name or not zelt:
        continue
    if vorname == "Vorname" or name == "Name":
        continue

    zelt_str = str(zelt).strip()
    key = (vorname, name)
    zelt_data[key] = zelt_str

print(f"  {len(zelt_data)} Zelt-Zuordnungen")

# ─── Zusammenführen & Importieren ───
print("\n📥 Importiere Teilnehmer...")
importiert = 0
uebersprungen = 0
ohne_zelt = 0

for (vorname, nachname), data in formular_data.items():
    # Zelt suchen (mit Toleranz bei Leerzeichen)
    zelt = zelt_data.get((vorname, nachname))
    if not zelt:
        # Versuch mit getrimmten Namen
        for (zv, zn), z in zelt_data.items():
            if zv.strip().lower() == vorname.strip().lower() and zn.strip().lower() == nachname.strip().lower():
                zelt = z
                break

    if not zelt:
        print(f"  ⚠️  Kein Zelt für {vorname} {nachname}")
        ohne_zelt += 1
        continue

    # Prüfen ob bereits vorhanden
    existing = conn.execute(
        "SELECT id FROM teilnehmer WHERE vorname = ? AND nachname = ?",
        (data["vorname"], data["nachname"])
    ).fetchone()

    if existing:
        print(f"  ⏭️  Bereits vorhanden: {vorname} {nachname}")
        uebersprungen += 1
        continue

    try:
        conn.execute(
            """INSERT INTO teilnehmer
               (vorname, nachname, zelt, anrede, kommentar, schwimmen, medikamente, hauptsportart)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (data["vorname"], data["nachname"], zelt,
             data["anrede"], data["kommentar"], data["schwimmen"],
             data["medikamente"], data["hauptsportart"])
        )
        importiert += 1
    except Exception as e:
        print(f"  ❌ Fehler bei {vorname} {nachname}: {e}")

conn.commit()
conn.close()

print(f"\n{'='*40}")
print(f"✅ Fertig!")
print(f"  Importiert:     {importiert}")
print(f"  Übersprungen:   {uebersprungen}")
print(f"  Ohne Zelt:      {ohne_zelt}")
print(f"{'='*40}")
