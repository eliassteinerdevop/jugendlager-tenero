#!/usr/bin/env python3
"""Fügt geburtsdatum-Spalte hinzu und importiert Geburtsdaten aus Excel."""

import sqlite3
import os
import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), "daten", "daten.db")
XLSX_PATH = os.path.join(os.path.dirname(__file__), "TN-Liste_Tenero 2026.xlsx")

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row

# Migration
cols = {r["name"] for r in conn.execute("PRAGMA table_info(teilnehmer)")}
if "geburtsdatum" not in cols:
    conn.execute("ALTER TABLE teilnehmer ADD COLUMN geburtsdatum TEXT NOT NULL DEFAULT ''")
    print("✅ Spalte 'geburtsdatum' hinzugefügt")
else:
    print("⏭️  Spalte 'geburtsdatum' existiert bereits")

# Geburtsdaten aus Zelt-Sheet
wb = openpyxl.load_workbook(XLSX_PATH)
ws_z = wb['Einteilung Zelte']

importiert = 0
for r in range(6, ws_z.max_row + 1):
    vorname = (ws_z.cell(row=r, column=2).value or "").strip()
    name = (ws_z.cell(row=r, column=3).value or "").strip()
    geburt = ws_z.cell(row=r, column=4).value

    if not vorname or not name or vorname == "Vorname" or not geburt:
        continue

    # Geburtsdatum als String normalisieren
    geb_str = str(geburt).strip()
    # Datetime-Objekte in String umwandeln
    if hasattr(geburt, 'strftime'):
        geb_str = geburt.strftime('%d.%m.%Y')

    # In DB updaten – Match über Vorname+Nachname (exakt und fuzzy)
    updated = conn.execute(
        "UPDATE teilnehmer SET geburtsdatum = ? WHERE vorname = ? AND nachname = ? AND geburtsdatum = ''",
        (geb_str, vorname, name)
    ).rowcount

    if updated == 0:
        # Fuzzy: Namen mit Leerzeichen vergleichen
        for tn in conn.execute("SELECT id, vorname, nachname FROM teilnehmer WHERE geburtsdatum = ''").fetchall():
            tv = tn["vorname"].strip().lower()
            tnn = tn["nachname"].strip().lower()
            ev = vorname.lower()
            en = name.lower()
            if (tv in ev or ev in tv) and (tnn in en or en in tnn):
                conn.execute("UPDATE teilnehmer SET geburtsdatum = ? WHERE id = ?", (geb_str, tn["id"]))
                updated = 1
                break

    if updated:
        importiert += 1

conn.commit()

# Prüfen ob alle ein Geburtsdatum haben
ohne = conn.execute("SELECT COUNT(*) FROM teilnehmer WHERE geburtsdatum = ''").fetchone()[0]
total = conn.execute("SELECT COUNT(*) FROM teilnehmer").fetchone()[0]
print(f"✅ Geburtsdaten importiert: {importiert}")
print(f"📊 Total: {total}, ohne Geburtsdatum: {ohne}")

if ohne > 0:
    print("\n⚠️  Ohne Geburtsdatum:")
    for r in conn.execute("SELECT vorname, nachname, zelt FROM teilnehmer WHERE geburtsdatum = ''").fetchall():
        print(f"   {r['vorname']} {r['nachname']}, Zelt {r['zelt']}")

conn.close()
